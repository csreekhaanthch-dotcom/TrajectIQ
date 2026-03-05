"""
TrajectIQ Enterprise - Data Exporters
=====================================
Export analytics data in multiple formats.
"""

import csv
import json
import io
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from enum import Enum


class ExportFormat(Enum):
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"


@dataclass
class ExportResult:
    success: bool
    format: ExportFormat
    content: Optional[bytes] = None
    file_path: Optional[str] = None
    row_count: int = 0
    error: Optional[str] = None


class DataExporter:
    """Exports analytics data in various formats."""

    def __init__(self, output_dir: str = None):
        self._logger = logging.getLogger(__name__)

    def export(self, data: Any, format: ExportFormat, filename: str = None, options: Dict = None) -> ExportResult:
        options = options or {}
        try:
            if format == ExportFormat.JSON:
                return self._export_json(data, filename, options)
            elif format == ExportFormat.CSV:
                return self._export_csv(data, filename, options)
            elif format == ExportFormat.EXCEL:
                return self._export_excel(data, filename, options)
            elif format == ExportFormat.PDF:
                return self._export_pdf(data, filename, options)
            return ExportResult(success=False, format=format, error=f"Unsupported format: {format}")
        except Exception as e:
            self._logger.error(f"Export failed: {e}")
            return ExportResult(success=False, format=format, error=str(e))

    def _export_json(self, data: Any, filename: str = None, options: Dict = None) -> ExportResult:
        indent = (options or {}).get("indent", 2)
        content = json.dumps(data, indent=indent, default=str).encode('utf-8')
        return ExportResult(success=True, format=ExportFormat.JSON, content=content, row_count=self._count_rows(data))

    def _export_csv(self, data: Any, filename: str = None, options: Dict = None) -> ExportResult:
        rows = self._normalize_to_rows(data)
        if not rows:
            return ExportResult(success=False, format=ExportFormat.CSV, error="No data to export")
        
        headers = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
        
        return ExportResult(success=True, format=ExportFormat.CSV, content=output.getvalue().encode('utf-8'), row_count=len(rows))

    def _export_excel(self, data: Any, filename: str = None, options: Dict = None) -> ExportResult:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self._logger.warning("openpyxl not available, falling back to CSV")
            return self._export_csv(data, filename, options)
        
        rows = self._normalize_to_rows(data)
        if not rows:
            return ExportResult(success=False, format=ExportFormat.EXCEL, error="No data to export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        headers = list(rows[0].keys()) if rows else []
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7c83fd", end_color="7c83fd", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, row_data in enumerate(rows, 2):
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        return ExportResult(success=True, format=ExportFormat.EXCEL, content=output.getvalue(), row_count=len(rows))

    def _export_pdf(self, data: Any, filename: str = None, options: Dict = None) -> ExportResult:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError:
            self._logger.warning("reportlab not available, falling back to JSON")
            return self._export_json(data, filename, options)
        
        rows = self._normalize_to_rows(data)
        if not rows:
            return ExportResult(success=False, format=ExportFormat.PDF, error="No data to export")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        
        headers = list(rows[0].keys()) if rows else []
        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in rows]
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c83fd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        doc.build([table])
        return ExportResult(success=True, format=ExportFormat.PDF, content=output.getvalue(), row_count=len(rows))

    def _normalize_to_rows(self, data: Any) -> List[Dict]:
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            for key in ['data', 'rows', 'items', 'results']:
                if key in data and isinstance(data[key], list):
                    return data[key]
            return [data]
        return [{"value": str(data)}]

    def _count_rows(self, data: Any) -> int:
        if isinstance(data, list):
            return len(data)
        return 1

    def export_candidates(self, candidates: List[Dict], format: ExportFormat = ExportFormat.CSV,
                         include_scores: bool = True) -> ExportResult:
        data = []
        for c in candidates:
            row = {"ID": c.get("candidate_id"), "Name": c.get("name"), "Email": c.get("email"),
                   "Status": c.get("status"), "Source": c.get("source")}
            if include_scores:
                scores = c.get("scores", {})
                row.update({"Overall": scores.get("overall"), "SDI": scores.get("sdi"),
                           "CSIG": scores.get("csig"), "IAE": scores.get("iae")})
            data.append(row)
        return self.export(data, format, f"candidates_{datetime.utcnow().strftime('%Y%m%d')}")

    def export_analytics(self, analytics_data: Dict, format: ExportFormat = ExportFormat.JSON,
                        report_name: str = "analytics") -> ExportResult:
        return self.export(analytics_data, format, f"{report_name}_{datetime.utcnow().strftime('%Y%m%d')}")
